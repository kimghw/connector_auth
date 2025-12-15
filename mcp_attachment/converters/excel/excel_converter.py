"""Excel to text converter."""

from pathlib import Path
import logging
from typing import List

from ...base_converter import BaseConverter

logger = logging.getLogger(__name__)


class ExcelConverter(BaseConverter):
    """Convert Excel files to text."""

    def convert(self, file_path: str) -> str:
        """Convert Excel file to text.

        Args:
            file_path: Path to the Excel file

        Returns:
            Extracted text content with sheet names and data
        """
        if not self.supports(file_path):
            raise ValueError(f"File {file_path} is not an Excel file")

        try:
            import openpyxl
            import pandas as pd

            # Use pandas for better data extraction
            excel_file = pd.ExcelFile(file_path)
            text_content = []

            for sheet_name in excel_file.sheet_names:
                df = excel_file.parse(sheet_name)

                if df.empty:
                    continue

                # Add sheet header
                text_content.append(f"=== Sheet: {sheet_name} ===")

                # Convert dataframe to readable text
                # Include column headers and data
                text_content.append(df.to_string(index=False, na_rep=''))
                text_content.append("")  # Add blank line between sheets

            return '\n'.join(text_content)

        except ImportError:
            # Fallback to openpyxl only
            try:
                return self._convert_with_openpyxl(file_path)
            except ImportError:
                logger.error("Neither pandas nor openpyxl installed.")
                logger.info("Install with: pip install pandas openpyxl")
                raise
        except Exception as e:
            logger.error(f"Error converting Excel {file_path}: {e}")
            raise

    def _convert_with_openpyxl(self, file_path: str) -> str:
        """Convert using openpyxl only.

        Args:
            file_path: Path to the Excel file

        Returns:
            Extracted text content
        """
        import openpyxl

        workbook = openpyxl.load_workbook(file_path, data_only=True)
        text_content = []

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text_content.append(f"=== Sheet: {sheet_name} ===")

            for row in sheet.iter_rows(values_only=True):
                # Filter out empty rows
                row_values = [str(cell) if cell is not None else '' for cell in row]
                if any(row_values):
                    text_content.append('\t'.join(row_values))

            text_content.append("")  # Add blank line between sheets

        workbook.close()
        return '\n'.join(text_content)

    def supports(self, file_path: str) -> bool:
        """Check if file is an Excel file.

        Args:
            file_path: Path to check

        Returns:
            True if file has Excel extension
        """
        path = Path(file_path)
        return path.suffix.lower() in ['.xlsx', '.xls', '.xlsm', '.xlsb']